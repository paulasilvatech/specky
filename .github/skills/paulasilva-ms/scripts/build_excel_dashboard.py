#!/usr/bin/env python3
"""Build a paulasilva-ms Office Excel dashboard, non-destructively.

This is the bundled helper for the Excel module of the paulasilva-ms skill
(references/excel.md). Standard library plus openpyxl only, no sandbox paths.

Use it two ways:

1. As a library (recommended): import the primitives and lay out a dashboard
   tailored to the real cells of a workbook. KPI values must be FORMULAS that
   reference existing cells, never retyped numbers. Snapshot the workbook
   first and verify nothing changed after saving.

2. As a CLI smoke test:
       python3 build_excel_dashboard.py input.xlsx output.xlsx
   This restyles every sheet to the Office standard, adds an empty
   'Dashboard' sheet, sets workbook properties, and asserts that no existing
   cell value changed (exits non-zero on any data drift).

Data integrity is the first rule: audited numbers are never fabricated or
retyped. The dashboard references them by formula.
"""
import sys
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    sys.exit(
        f"ERROR: a required dependency is missing ({exc.name}).\n"
        "  pip install -r scripts/requirements.txt\n"
        "  (this script needs openpyxl)"
    )

# ---- paulasilva-ms Office tokens (ARGB) ----
FONT = 'Segoe UI'
MONO = 'Consolas'
INK = 'FF201F1E'
BLUE = 'FF0078D4'
GRAY = 'FF605E5C'
ZEBRA = 'FFF3F2F1'
WHITE = 'FFFFFFFF'
LINE = 'FFE1DFDD'
ACCENTS = ['FF0078D4', 'FFF25022', 'FF107C10', 'FFFFB900', 'FF038387', 'FF5C2D91']
THIN = Side(style='thin', color=LINE)
WBORD = Side(style='thin', color='FFFFFFFF')


# ---- data-integrity pair ----
def snapshot(wb):
    """Capture every non-empty cell value as {(sheet, coord): value}."""
    snap = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    snap[(ws.title, c.coordinate)] = c.value
    return snap


def verify_unchanged(path, snap, tol=1e-6):
    """Reload path and assert every snapshotted cell is unchanged.

    Float values are compared with a small tolerance so that float repr noise
    (1459.8709000000001 vs 1459.8709) is not treated as a data change.
    Returns the count of genuine changes (0 means clean); prints each change.
    """
    wb = openpyxl.load_workbook(path)
    changed = 0
    for (sh, co), v in snap.items():
        nv = wb[sh][co].value
        same = (nv == v) or (
            isinstance(nv, (int, float)) and isinstance(v, (int, float))
            and not isinstance(nv, bool) and not isinstance(v, bool)
            and abs(nv - v) <= tol
        )
        if not same:
            changed += 1
            print(f"   CHANGED {sh}!{co}: {v!r} -> {nv!r}")
    return changed


# ---- restyle existing sheets ----
def _font(c, size=None, bold=None, color=None):
    f = c.font
    return Font(name=FONT, size=size or f.size or 11,
                bold=f.bold if bold is None else bold,
                italic=f.italic, color=color or (f.color if f.color else None))


def restyle(ws, tab=None):
    """Apply the Office look to an existing sheet, preserving all values."""
    ws.sheet_view.showGridLines = False
    if tab:
        ws.sheet_properties.tabColor = tab
    maxr, maxc = ws.max_row, ws.max_column
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None or (c.font and c.font.name):
                c.font = _font(c)

    def filled(r):
        return [c.column for c in ws[r] if c.value is not None]

    hdr, cols = None, None
    for r in range(1, min(13, maxr) + 1):
        f = filled(r)
        if len(f) >= 4 and r < maxr and len(filled(r + 1)) >= 3:
            hdr, cols = r, f
            break
    if hdr:
        minc, maxc2 = min(cols), max(cols)
        for cc in range(minc, maxc2 + 1):
            c = ws.cell(row=hdr, column=cc)
            c.fill = PatternFill('solid', fgColor=BLUE)
            c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = Border(bottom=WBORD, right=WBORD)
        ws.row_dimensions[hdr].height = 24
        for i, r in enumerate(range(hdr + 1, maxr + 1)):
            for cc in range(minc, maxc2 + 1):
                c = ws.cell(row=r, column=cc)
                if i % 2 == 1 and c.fill.patternType is None:
                    c.fill = PatternFill('solid', fgColor=ZEBRA)
                if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                    if abs(c.value) >= 1000 and c.number_format in ('General', '@'):
                        c.number_format = '#,##0'
        ws.freeze_panes = ws.cell(row=hdr + 1, column=minc)
        ws.auto_filter.ref = (
            f"{get_column_letter(minc)}{hdr}:{get_column_letter(maxc2)}{maxr}")
    for cc in range(1, maxc + 1):
        best = 0
        for r in range(1, maxr + 1):
            v = ws.cell(row=r, column=cc).value
            if v is not None:
                best = max(best, len(str(v)))
        if best:
            ws.column_dimensions[get_column_letter(cc)].width = min(max(best + 2, 10), 46)


def set_props(wb, title, desc, subject='GitHub Copilot Usage-Based Billing',
              keywords='GitHub Copilot, Usage-Based Billing'):
    p = wb.properties
    p.creator = 'Paula Silva, Software Global Black Belt'
    p.lastModifiedBy = 'Paula Silva, Software Global Black Belt'
    p.title = title
    p.subject = subject
    p.keywords = keywords
    p.description = desc
    p.category = 'GitHub Copilot UBB'
    try:
        p.company = 'Microsoft'
    except Exception:
        pass


# ---- dashboard primitives ----
def dash_base(ws, eyebrow, title, subtitle):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE
    ws.column_dimensions['A'].width = 2.4
    for L in 'BCDEFGHI':
        ws.column_dimensions[L].width = 16.5
    ws['B2'] = eyebrow
    ws['B2'].font = Font(FONT, 9, bold=True, color=BLUE)
    ws['B3'] = title
    ws['B3'].font = Font(FONT, 26, bold=True, color=INK)
    ws['B4'] = subtitle
    ws['B4'].font = Font(FONT, 11, color=GRAY)
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[4].height = 20


def kpi(ws, col, row, label, formula, caption, accent, numfmt='#,##0'):
    """One KPI card. `formula` must reference existing cells, e.g. =Sheet!C9."""
    for rr in (row, row + 1, row + 2):
        ws.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=col + 1)
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = PatternFill('solid', fgColor=accent)
    lc.font = Font(FONT, 9, bold=True, color=WHITE)
    lc.alignment = Alignment('left', 'center', indent=1)
    vc = ws.cell(row=row + 1, column=col, value=formula)
    vc.font = Font(FONT, 20, bold=True, color=INK)
    vc.fill = PatternFill('solid', fgColor=WHITE)
    vc.alignment = Alignment('left', 'center', indent=1)
    vc.number_format = numfmt
    cc = ws.cell(row=row + 2, column=col, value=caption)
    cc.font = Font(FONT, 9, color=GRAY)
    cc.fill = PatternFill('solid', fgColor=WHITE)
    cc.alignment = Alignment('left', 'top', indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 30
    ws.row_dimensions[row + 2].height = 26
    for rr in (row, row + 1, row + 2):
        for k in (col, col + 1):
            ws.cell(row=rr, column=k).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def cards_band(ws, row, items):
    """A band of up to three KPI cards. items: [(label, formula, caption, numfmt), ...]."""
    for i, (label, formula, caption, fmt) in enumerate(items):
        kpi(ws, 2 + i * 3, row, label, formula, caption, ACCENTS[i % len(ACCENTS)], fmt)


def section(ws, row, label):
    ws.cell(row=row, column=2, value=label).font = Font(FONT, 12, bold=True, color=INK)
    ws.row_dimensions[row].height = 22


def bullets(ws, row, lines):
    for i, t in enumerate(lines):
        r = row + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        c = ws.cell(row=r, column=2, value="\u2022  " + t)
        c.font = Font(FONT, 10, color=INK if i == 0 else GRAY)
        ws.row_dimensions[r].height = 18


def footer(ws, row, text):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    c = ws.cell(row=row, column=2, value=text)
    c.font = Font(FONT, 8.5, italic=True, color=GRAY)


# ---- chart helpers (all reference existing ranges) ----
def bar(ws, title, cat_ref, val_ref, anchor, width=15.5, height=8.2, color='FF0078D4'):
    ch = BarChart()
    ch.type = 'bar'
    ch.title = title
    ch.style = 2
    ch.height, ch.width = height, width
    ch.add_data(val_ref, titles_from_data=True)
    ch.set_categories(cat_ref)
    ch.legend = None
    try:
        ch.series[0].graphicalProperties.solidFill = color[2:]
    except Exception:
        pass
    ws.add_chart(ch, anchor)


def colchart(ws, title, cat_ref, val_ref, anchor, width=11.5, height=8.2, color='FFF25022'):
    ch = BarChart()
    ch.type = 'col'
    ch.title = title
    ch.style = 2
    ch.height, ch.width = height, width
    ch.add_data(val_ref, titles_from_data=True)
    ch.set_categories(cat_ref)
    ch.legend = None
    try:
        ch.series[0].graphicalProperties.solidFill = color[2:]
    except Exception:
        pass
    ws.add_chart(ch, anchor)


def line(ws, title, cat_ref, val_ref, anchor, width=20.0, height=8.5, color='FF0078D4'):
    ch = LineChart()
    ch.title = title
    ch.style = 2
    ch.height, ch.width = height, width
    ch.add_data(val_ref, titles_from_data=True)
    ch.set_categories(cat_ref)
    ch.legend = None
    try:
        ch.series[0].graphicalProperties.line.solidFill = color[2:]
        ch.series[0].graphicalProperties.line.width = 28000
    except Exception:
        pass
    ws.add_chart(ch, anchor)


# ---- CLI smoke test ----
def _cli(src, dst):
    wb = openpyxl.load_workbook(src)
    snap = snapshot(wb)
    existing = list(wb.sheetnames)
    name = 'Dashboard'
    if name in wb.sheetnames:
        name = 'Dashboard (paulasilva-ms)'
    ws = wb.create_sheet(name, 0)
    dash_base(ws, 'PAULASILVA-MS  ·  OFFICE STANDARD',
              'Dashboard', 'Smoke test. Replace with KPI cards and charts that '
              'reference the detail sheets.')
    section(ws, 7, 'How to read')
    bullets(ws, 8, [
        'This empty dashboard was added by the CLI smoke test.',
        'Use the library primitives to add KPI cards (formulas) and native charts.',
        'Every KPI value must reference an existing cell, never a retyped number.',
    ])
    footer(ws, 12, 'Author: Paula Silva, Software Global Black Belt, Microsoft.')
    for nm in existing:
        restyle(wb[nm])
    set_props(wb, 'paulasilva-ms Excel dashboard', 'Smoke test output.')
    wb.save(dst)
    changed = verify_unchanged(dst, snap)
    print(f"OK: sheets {existing} + Dashboard; cells preserved={len(snap)} changed={changed}")
    return changed


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print('usage: python3 build_excel_dashboard.py input.xlsx output.xlsx')
        sys.exit(2)
    sys.exit(1 if _cli(sys.argv[1], sys.argv[2]) > 0 else 0)
